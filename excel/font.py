from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from copy import copy

book = Workbook()
sheet = book.active

rows = (
    (34, 26),
    (88, 36),
    (24, 29),
    (15, 22),
    (56, 13),
    (76, 18)
)

for row in rows:
    sheet.append(row)
    
font = Font(name="Arial",size=14,bold=True,color='FF0000')
cell = sheet.cell(row=7, column=2)
cell.value = "=SUM(A1:B6)"
cell.font = copy(font)
cell.fill = PatternFill(patternType="gradient",start_color='FF00FF',end_color='FFEE22',fill_type="lightUp")

book.save('xlf/8.xlsx')