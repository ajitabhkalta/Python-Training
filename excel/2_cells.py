import openpyxl
from openpyxl.styles import Alignment

book = openpyxl.load_workbook('xlf/7.xlsx')

sheet = book["Sheet1"]
sheet.sheet_properties.tabColor = "00ff33"
book.save('xlf/7.xlsx')
sheet = book.active

sheet.merge_cells('A1:B2')

cell = sheet.cell(row=1, column=1)
cell.value = 'WOW ITS BIG'
cell.alignment = Alignment(horizontal='left', vertical='center')

book.save('xlf/7.xlsx')
