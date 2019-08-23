import openpyxl

book = openpyxl.Workbook()
book.save("xlf/6.xlsx")

book = openpyxl.load_workbook('xlf/6.xlsx')
book.create_sheet("Qrtr-3")
book.create_sheet("Qrtr-2")

print(book.sheetnames)
book.save('xlf/6.xlsx')
input()
sheet1 = book["Qrtr-2"]
book.remove(sheet1)
book.save('xlf/6.xlsx')
input()
print(book.sheetnames)
book.create_sheet("Qrtr-1", 1)
#save data
sheet=book["Qrtr-3"]
sheet["A1"]="PAVAN"
book.save('xlf/6.xlsx')
print(book.sheetnames)